# -*- coding: utf-8 -*-
import streamlit as st
import streamlit.components.v1 as components
import math, json, os, re, requests, time, datetime, base64

# ── GitHub 연동 함수 ──────────────────────────────────────
GITHUB_TOKEN = st.secrets.get("GITHUB_TOKEN", "")
GITHUB_REPO  = st.secrets.get("GITHUB_REPO", "noah-krw/ops-manager-tool")
GITHUB_FILE  = st.secrets.get("GITHUB_FILE", "merchants.json")
GITHUB_API   = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"

def load_data():
    """GitHub에서 merchants.json 읽기"""
    try:
        headers = {"Authorization": f"token {GITHUB_TOKEN}",
                   "Accept": "application/vnd.github+json"}
        r = requests.get(GITHUB_API, headers=headers, timeout=5)
        if r.status_code == 200:
            content = base64.b64decode(r.json()["content"]).decode("utf-8")
            data = json.loads(content)
            if "merchants" not in data:
                return {"my_wallet": data.get("my_wallet", ""), "merchants": data}
            return data
        else:
            st.warning(f"⚠️ GitHub 읽기 실패 ({r.status_code}) - 기본값으로 시작합니다.")
    except Exception as e:
        st.warning(f"⚠️ GitHub 연결 오류: {e}")
    return get_default_data()

def save_data(data):
    """GitHub에 merchants.json 저장"""
    try:
        headers = {"Authorization": f"token {GITHUB_TOKEN}",
                   "Accept": "application/vnd.github+json"}
        # 현재 파일 SHA 가져오기 (업데이트에 필요)
        r = requests.get(GITHUB_API, headers=headers, timeout=5)
        if r.status_code != 200:
            st.error(f"GitHub 읽기 실패: {r.status_code} {r.text}")
            return
        sha = r.json().get("sha", "")
        content = base64.b64encode(json.dumps(data, indent=4, ensure_ascii=False).encode()).decode()
        payload = {
            "message": "Update merchants.json via Streamlit",
            "content": content,
            "sha": sha
        }
        res = requests.put(GITHUB_API, headers=headers, json=payload, timeout=5)
        if res.status_code in [200, 201]:
            st.toast("GitHub에 저장 완료 ✅")
        else:
            st.error(f"GitHub 저장 실패: {res.status_code} {res.text}")
    except Exception as e:
        st.error(f"저장 실패: {e}")

def get_default_data():
    return {
        'my_wallet': 'TDaQt8oASZhVsuaEdpevqCacGKseGKCWhQ',
        'merchants': {
            'spfxm': {'wallet': 'TWbFbzW5GRkAkVrY4fLuXhNA576DCkoGbh', 'fee': '4', 'note': '메인 업체'}
        }
    }

def extract_int(text):
    if not text: return 0
    num_str = re.sub(r'[^0-9]', '', str(text))
    return int(num_str) if num_str else 0

def fmt(n): return f"{n:,}"

def editable_box(text, color_type="blue", box_id="default"):
    if not text: return
    colors = {
        "blue":   {"border": "#4a90d9", "glow": "rgba(74,144,217,0.25)", "bg": "#07101e", "text": "#a8c7e8", "btn_bg": "#0d2040", "btn_hover": "#4a90d9"},
        "green":  {"border": "#2ecc71", "glow": "rgba(46,204,113,0.2)",  "bg": "#04120a", "text": "#7dcea0", "btn_bg": "#0a2016", "btn_hover": "#2ecc71"},
        "yellow": {"border": "#f39c12", "glow": "rgba(243,156,18,0.2)",  "bg": "#16100a", "text": "#f8c471", "btn_bg": "#241800", "btn_hover": "#f39c12"},
        "red":    {"border": "#e74c3c", "glow": "rgba(231,76,60,0.2)",   "bg": "#160608", "text": "#f1948a", "btn_bg": "#240808", "btn_hover": "#e74c3c"},
        "sky":    {"border": "#5dade2", "glow": "rgba(93,173,226,0.2)",  "bg": "#081622", "text": "#5dade2", "btn_bg": "#0c1e30", "btn_hover": "#5dade2"},
    }
    c = colors.get(color_type, colors["blue"])
    line_count = text.count("\n") + 1
    height = max(160, line_count * 26 + 90)
    html_code = f"""
    <div style="margin-bottom:14px;">
        <div style="border-left:3px solid {c['border']};border-radius:0 8px 8px 0;
            box-shadow:0 0 18px {c['glow']},inset 0 0 30px rgba(0,0,0,0.3);
            overflow:hidden;background:{c['bg']};">
            <textarea id="copy_area_{box_id}" style="width:100%;height:{height-55}px;
                background:#ffffff;color:#1a1a1a;border:none;outline:none;
                font-family:'Courier New',monospace;font-size:14px;line-height:1.7;
                padding:14px 16px;box-sizing:border-box;resize:none;">{text}</textarea>
            <div style="display:flex;align-items:center;justify-content:flex-end;
                gap:10px;padding:8px 12px 10px;border-top:1px solid rgba(255,255,255,0.08);
                background:rgba(0,0,0,0.2);">
                <span style="font-family:'Courier New',monospace;font-size:13px;
                    color:{c['text']};opacity:0.6;">✎ 직접 수정 가능</span>
                <button id="btn_{box_id}" onclick="copyText_{box_id}()" style="
                    padding:7px 18px;background:{c['btn_bg']};color:{c['text']};
                    border:1px solid {c['border']};border-radius:6px;cursor:pointer;
                    font-family:'Courier New',monospace;font-size:13px;font-weight:600;">COPY</button>
            </div>
        </div>
    </div>
    <script>
    function copyText_{box_id}() {{
        const t = document.getElementById('copy_area_{box_id}');
        const b = document.getElementById('btn_{box_id}');
        t.select(); t.setSelectionRange(0,99999);
        try {{
            document.execCommand('copy');
            b.innerText='COPIED'; b.style.background='{c['border']}'; b.style.color='#000';
            setTimeout(()=>{{b.innerText='COPY';b.style.background='{c['btn_bg']}';b.style.color='{c['text']}';}},1800);
        }} catch(e) {{}}
    }}
    </script>"""
    components.html(html_code, height=height)


st.set_page_config(page_title="단계별 정산 시스템 v97", layout="centered")
st.markdown("""
<link href="https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Noto+Sans+KR:wght@300;400;500;700&display=swap" rel="stylesheet">
<style>
    [data-testid="stAppViewContainer"] {
        background-color: #060c16 !important;
        background-image: radial-gradient(ellipse at 20% 0%, rgba(30,60,100,0.35) 0%, transparent 60%),
            radial-gradient(ellipse at 80% 100%, rgba(20,80,60,0.2) 0%, transparent 60%);
        color: #c8d6e5 !important;
    }
    [data-testid="stHeader"] { background: transparent !important; }
    [data-testid="stSidebar"] { background: #080e1a !important; border-right: 1px solid rgba(74,144,217,0.15) !important; }
    [data-testid="stSidebar"] button {
        width:100%; background:#0d1a2e !important; color:#7aa8cc !important;
        border:1px solid rgba(74,144,217,0.25) !important; border-radius:8px !important;
        font-family:'Space Mono',monospace !important; font-size:0.9em !important;
        font-weight:600 !important; letter-spacing:0.08em !important;
        padding:11px 0 !important; transition:all 0.2s ease !important;
    }
    [data-testid="stSidebar"] button:hover {
        background:#1a2e48 !important; border-color:#4a90d9 !important;
        color:#cde4f8 !important; box-shadow:0 0 12px rgba(74,144,217,0.2) !important;
    }
    .main-title {
        font-family:'Space Mono',monospace; color:#4a90d9; font-size:2.4em; font-weight:700;
        text-align:center; letter-spacing:0.22em; text-transform:uppercase;
        margin-bottom:28px; padding:22px 0 18px;
        border-bottom:1px solid rgba(74,144,217,0.25);
        text-shadow:0 0 30px rgba(74,144,217,0.4);
    }
    .main-title::after { content:''; display:block; width:80px; height:2px;
        background:linear-gradient(90deg,transparent,#4a90d9,transparent); margin:12px auto 0; }
    .section-header {
        display:flex; align-items:center; gap:12px; margin-top:32px; margin-bottom:14px;
        padding:10px 16px; background:rgba(255,255,255,0.02); border-radius:8px;
        border-left:3px solid var(--hdr-color,#4a90d9); position:relative; overflow:hidden;
    }
    .section-header::before { content:''; position:absolute; top:0;left:0;right:0;bottom:0;
        background:linear-gradient(90deg,rgba(var(--hdr-rgb,74,144,217),0.08) 0%,transparent 100%); pointer-events:none; }
    .section-header .num { font-family:'Space Mono',monospace; font-size:0.75em; font-weight:700;
        color:var(--hdr-color,#4a90d9); letter-spacing:0.1em; opacity:0.7; }
    .section-header .title { font-family:'Noto Sans KR',sans-serif; font-size:0.92em; font-weight:700;
        color:#d0dff0; letter-spacing:0.08em; text-transform:uppercase; }
    .section-header .line { flex:1; height:1px;
        background:linear-gradient(90deg,rgba(var(--hdr-rgb,74,144,217),0.3) 0%,transparent 100%); }
    div[data-baseweb="input"] { background-color:#0b1525 !important; border-radius:7px !important;
        border:1px solid rgba(74,144,217,0.25) !important; }
    div[data-baseweb="input"]:focus-within { border-color:rgba(74,144,217,0.7) !important; }
    input { color:#dce8f5 !important; font-family:'Space Mono',monospace !important;
        font-size:1.1em !important; background:transparent !important; caret-color:#4a90d9 !important; }
    label[data-testid="stWidgetLabel"] > div > p { color:#6a8aaa !important; font-size:0.82em !important;
        font-family:'Space Mono',monospace !important; letter-spacing:0.06em !important; text-transform:uppercase !important; }
    [data-testid="stRadio"] label { font-family:'Space Mono',monospace !important; font-size:0.88em !important; color:#7aa8cc !important; }
    div[data-baseweb="select"] > div { background-color:#0b1525 !important;
        border:1px solid rgba(74,144,217,0.25) !important; border-radius:7px !important;
        color:#a8c7e8 !important; font-family:'Space Mono',monospace !important; }
    .rate-text { font-family:'Space Mono',monospace; color:#5dade2; font-size:0.95em;
        margin:8px 0 6px; padding:8px 14px; background:rgba(93,173,226,0.07); border-radius:6px; }
    hr { border-color:rgba(74,144,217,0.1) !important; margin:24px 0 !important; }
    div[data-testid="stSidebar"] div:has(> button[key="reset_inputs"]) button {
        background:linear-gradient(135deg,#0d2a1a,#0a1f2e) !important;
        border-color:rgba(46,204,113,0.4) !important; color:#2ecc71 !important; }
    div[data-testid="stSidebar"] div:has(> button[key="reset_inputs"]) button:hover {
        background:#2ecc71 !important; color:#000 !important; }
    ::-webkit-scrollbar { width:5px; }
    ::-webkit-scrollbar-track { background:#060c16; }
    ::-webkit-scrollbar-thumb { background:#1e3a5f; border-radius:10px; }
</style>
""", unsafe_allow_html=True)

def section_header(num, title, color="#4a90d9", rgb="74,144,217"):
    st.markdown(f"""
    <div class="section-header" style="--hdr-color:{color};--hdr-rgb:{rgb};">
        <span class="num">{num}</span>
        <span class="title">{title}</span>
        <span class="line"></span>
    </div>""", unsafe_allow_html=True)

if 'db' not in st.session_state: st.session_state.db = load_data()
if 'page' not in st.session_state: st.session_state.page = 'settle'

if 'bithumb_price' not in st.session_state:
    try:
        _r = requests.get('https://api.bithumb.com/public/ticker/USDT_KRW', timeout=3)
        if _r.json().get('status') == '0000':
            st.session_state.bithumb_price = int(float(_r.json()['data']['closing_price']))
        else:
            st.session_state.bithumb_price = 0
    except:
        st.session_state.bithumb_price = 0
    st.session_state.bithumb_ts = time.time()

with st.sidebar:
    st.markdown("""
    <div style="font-family:'Space Mono',monospace;font-size:1.1em;font-weight:700;
        color:#4a90d9;letter-spacing:0.15em;text-align:center;
        padding:10px 0 20px;border-bottom:1px solid rgba(74,144,217,0.2);margin-bottom:16px;">
        💹 SETTLEMENT
    </div>""", unsafe_allow_html=True)
    if st.button("🚀  USDT 정산", use_container_width=True):
        st.session_state.page = 'settle'; st.rerun()
    if st.button("📤  USDT 탑업", use_container_width=True):
        st.session_state.page = 'topup'; st.rerun()
    if st.button("⚙️  머천트 관리", use_container_width=True):
        st.session_state.page = 'admin'; st.rerun()
    if st.button("📊  수수료 정산", use_container_width=True):
        st.session_state.page = 'commission'; st.rerun()
    if st.button("👤  에이전트 정산", use_container_width=True):
        st.session_state.page = 'agent'; st.rerun()
    st.divider()
    reset_keys = ["t_b","t_u","t_k","t_s"] if st.session_state.page == "topup" else ["s_b","s_s","s_amt","bal_in","w_in"]
    if st.button("⟳  NEW SESSION", key="reset_inputs", use_container_width=True):
        for k in reset_keys: st.session_state[k] = ""
        st.toast("입력값이 초기화되었습니다", icon="🔄"); st.rerun()
    st.divider()
    if st.button("🔄  데이터 복구", use_container_width=True):
        st.session_state.db = get_default_data()
        save_data(st.session_state.db); st.success("복구 완료"); st.rerun()

# ══════════════════════════════════════════════════════════
# 정산 페이지
# ══════════════════════════════════════════════════════════
if st.session_state.page == 'settle':
    st.markdown('<div class="main-title">USDT 정산</div>', unsafe_allow_html=True)
    merchants = st.session_state.db['merchants']
    sorted_keys = sorted(list(merchants.keys()))
    default_idx = sorted_keys.index('spfxm') if 'spfxm' in sorted_keys else 0
    selected_m = st.selectbox("", sorted_keys, index=default_idx, label_visibility="collapsed")
    m_info = merchants[selected_m]
    m_map = {"4%": 1.04, "4.5%": 1.045, "5%": 1.05}
    sel_p = st.radio("", ["4%", "4.5%", "5%"], index=0, horizontal=True, label_visibility="collapsed")

    section_header("01", "정산 환율", "#4a90d9", "74,144,217")
    live_price = st.session_state.get("bithumb_price", 0)
    kst = datetime.timezone(datetime.timedelta(hours=9))
    fetched_time = datetime.datetime.fromtimestamp(st.session_state.get("bithumb_ts", time.time()), tz=kst).strftime("%H:%M:%S")
    bithumb_str = ("&#8361; " + fmt(live_price)) if live_price > 0 else "&mdash;"
    html = (
        "<style>@keyframes blink{0%,100%{opacity:1;}50%{opacity:0.15;}}</style>"
        "<div style='padding:14px 22px;margin-bottom:14px;background:linear-gradient(135deg,#030f1c,#041810);"
        "border:1px solid rgba(93,173,226,0.3);border-radius:10px;box-shadow:0 0 18px rgba(93,173,226,0.1);"
        "display:flex;align-items:center;justify-content:space-between;'>"
        "<div style='display:flex;align-items:center;gap:8px;'>"
        "<span style='display:inline-block;width:7px;height:7px;border-radius:50%;background:#2ecc71;"
        "box-shadow:0 0 7px #2ecc71;animation:blink 1.5s infinite;'></span>"
        "<span style='font-family:Space Mono,monospace;font-size:0.68em;color:#5dade2;letter-spacing:0.1em;'>BITHUMB &nbsp; USDT / KRW</span>"
        "</div>"
        "<div style='display:flex;flex-direction:column;align-items:center;gap:3px;'>"
        "<div style='font-family:Space Mono,monospace;font-size:1.8em;font-weight:700;color:#ffffff;'>" + bithumb_str + "</div>"
        "<div style='font-family:Space Mono,monospace;font-size:0.68em;color:#5dade2;'>" + fetched_time + "</div>"
        "</div>"
        "<a href='https://search.naver.com/search.naver?query=%EB%B9%97%EC%8D%B8+%ED%85%8C%EB%8D%94+%EC%8B%9C%EC%84%B8' "
        "target='_blank' style='font-family:Space Mono,monospace;font-size:0.85em;font-weight:700;"
        "color:#2ecc71;border:1px solid rgba(46,204,113,0.5);border-radius:8px;padding:10px 20px;"
        "text-decoration:none;background:rgba(46,204,113,0.08);'>김프 확인</a></div>"
    )
    st.markdown(html, unsafe_allow_html=True)
    if st.button('⟳  시세 새로고침', key='refresh_ticker'):
        if 'bithumb_price' in st.session_state: del st.session_state['bithumb_price']
        st.rerun()

    sc1, sc2 = st.columns(2)
    with sc1:
        if st.session_state.get("s_b", "") == "" and live_price > 0:
            st.session_state["s_b"] = str(live_price)
        sb_val = extract_int(st.text_input("빗썸 시세", key="s_b"))
    with sc2: ss_val = extract_int(st.text_input("수동 환율", key="s_s"))
    s_rate = ss_val if ss_val > 0 else math.ceil(sb_val * m_map[sel_p])
    if s_rate > 0: editable_box(f"1usdt = {fmt(s_rate)} krw", "sky", "rate_01")

    section_header("02", "정산 멘트 생성", "#4a90d9", "74,144,217")
    amt = extract_int(st.text_input("정산 금액 (KRW) 입력", key="s_amt"))
    if amt > 0 and s_rate > 0:
        u_val = round(amt / s_rate, 2)
        s_msg = (f"mid : {selected_m}\nsettlement amount : {fmt(amt)} krw\n"
                 f"exchange to usdt : {u_val:,.2f} usdt\n1usdt = {fmt(s_rate)} krw\n\n"
                 f"{m_info['wallet']}\n\nPlease confirm the address and calculation.\n"
                 f"Once approved, we will proceed immediately")
        editable_box(s_msg, "blue", "res_02")

    section_header("03", "최종 잔액 보고", "#4a90d9", "74,144,217")
    bal_in = extract_int(st.text_input("현재 잔액 입력 (KRW)", key="bal_in"))
    if bal_in > 0 and amt > 0 and s_rate > 0:
        u_ceil = math.ceil(amt / s_rate)
        b_msg = (f"Balance & settlement update\n\n- {selected_m}\n"
                 f"settlement amount : {fmt(amt)} krw\nexchange to usdt : {fmt(u_ceil)} usdt\n"
                 f"1usdt = {fmt(s_rate)} krw\n\n- {selected_m} : {fmt(bal_in)} krw")
        editable_box(b_msg, "green", "res_03")

    section_header("04", "마크업 수수료", "#f39c12", "243,156,18")
    if amt > 0:
        m_fee = float(m_info.get('fee', 0.5))
        markup = math.ceil(amt * (m_fee / 100))
        markup_msg = f"드래곤 테더정산 마크업 {m_fee}% {selected_m} / {fmt(amt)} / {fmt(markup)}"
        editable_box(markup_msg, "yellow", "res_04")

    section_header("05", "정산 (SETTLEMENT) 요청", "#e74c3c", "231,76,60")
    w_bal = extract_int(st.text_input("하이 밸런스 경고용 잔액", key="w_in"))
    if w_bal > 0 and s_rate > 0:
        st.markdown(f'<p class="rate-text">▸ 적용 환율 &nbsp; 1usdt = {fmt(s_rate)} krw</p>', unsafe_allow_html=True)
        w_msg = (f"Hello, Team\nCurrently, the balance of the merchants is too high.\n"
                 f"To ensure a safe balance, please proceed with USDT settlement.\nThank you\n\n"
                 f"Balance update\n\n- {selected_m} : {fmt(w_bal)} krw")
        editable_box(w_msg, "red", "res_05")

# ══════════════════════════════════════════════════════════
# 탑업 페이지
# ══════════════════════════════════════════════════════════
elif st.session_state.page == 'topup':
    st.markdown('<div class="main-title">USDT 탑업</div>', unsafe_allow_html=True)
    merchants = st.session_state.db['merchants']
    sorted_keys = sorted(list(merchants.keys()))
    default_idx = sorted_keys.index('spfxm') if 'spfxm' in sorted_keys else 0
    selected_m = st.selectbox("", sorted_keys, index=default_idx, label_visibility="collapsed", key="topup_merchant")
    m_info = merchants[selected_m]
    live_price = st.session_state.get("bithumb_price", 0)
    kst = datetime.timezone(datetime.timedelta(hours=9))
    fetched_time = datetime.datetime.fromtimestamp(st.session_state.get("bithumb_ts", time.time()), tz=kst).strftime("%H:%M:%S")
    bithumb_str = ("&#8361; " + fmt(live_price)) if live_price > 0 else "&mdash;"
    st.markdown("""
    <div style='display:flex;align-items:center;gap:12px;margin-bottom:12px;'>
        <a href='https://www.google.com/search?q=테더+원화+시세&hl=ko&gl=KR'
           target='_blank' style='font-family:Space Mono,monospace;font-size:0.85em;font-weight:700;
           color:#f39c12;border:1px solid rgba(243,156,18,0.5);border-radius:8px;padding:10px 24px;
           text-decoration:none;background:rgba(243,156,18,0.08);'>🔍 구글 시세 확인</a>
    </div>
    <div style='padding:10px 16px;background:rgba(243,156,18,0.08);
        border-left:3px solid #f39c12;border-radius:6px;
        font-family:Noto Sans KR,sans-serif;font-size:0.88em;color:#f8c471;'>
        ⚠️ 탑업은 <b>구글 테더 시세</b> 기준으로 진행합니다. 위 버튼을 확인 후 아래에 입력하세요.
    </div>
    """, unsafe_allow_html=True)
    section_header("01", "TOP-UP 탑업", "#2ecc71", "46,204,113")

    with st.expander("📋 신규 업체 탑업 안내 (클릭하여 펼치기)"):
        notice_msg = ("merchant top-up notice\n\n"
                      "@ Please let us know the amount when request top-up.\n"
                      "@ Please send USDT as soon as possible after top up approval.\n"
                      "  If processing is delayed, provided exchange rate may change\n"
                      "  based on the real-time exchange rate again.\n"
                      "@ After top-up approval cannot be cancelled.\n\n"
                      "thank you.")
        editable_box(notice_msg, "sky", "res_notice")
    ts_val = extract_int(st.text_input("구글 테더 시세 입력 (KRW)", key="t_s",
                                        placeholder="구글에서 확인한 테더 시세를 입력하세요"))
    tb_val = live_price  # 참고용

    if ts_val > 0:
        rate_05 = math.floor(ts_val * (1 - 0.005))  # -0.5%
        rate_1  = math.floor(ts_val * (1 - 0.01))   # -1%
        st.markdown(
            f"<div style='display:flex;gap:12px;margin:10px 0;font-family:Space Mono,monospace;font-size:0.9em;'>"
            f"<div style='padding:8px 16px;background:rgba(93,173,226,0.08);border:1px solid rgba(93,173,226,0.4);border-radius:8px;color:#5dade2;'>"
            f"-0.5% &nbsp; <b>{fmt(rate_05)} krw</b></div>"
            f"<div style='padding:8px 16px;background:rgba(46,204,113,0.08);border:1px solid rgba(46,204,113,0.4);border-radius:8px;color:#2ecc71;'>"
            f"-1% &nbsp; <b>{fmt(rate_1)} krw</b></div>"
            f"</div>",
            unsafe_allow_html=True
        )
        t_mode_rate = st.radio("", [f"-0.5% : {fmt(rate_05)} krw", f"-1% : {fmt(rate_1)} krw", "수동 입력"],
                               horizontal=True, label_visibility="collapsed", key="t_rate_mode")
        if "수동 입력" in t_mode_rate:
            t_rate_manual = extract_int(st.text_input("수동 환율 직접 입력 (KRW)", key="t_manual",
                                                       placeholder="관리자 지정 환율을 입력하세요"))
            t_rate = t_rate_manual if t_rate_manual > 0 else 0
        else:
            t_rate = rate_05 if "-0.5%" in t_mode_rate else rate_1
    else:
        t_rate = 0
    if t_rate > 0:
        st.markdown(f"<div style='font-family:Space Mono,monospace;color:#5dade2;font-size:0.95em;"
                    f"margin-bottom:12px;padding:8px 14px;background:rgba(93,173,226,0.07);border-radius:6px;'>"
                    f"▸ 적용 환율 &nbsp; 1usdt = {fmt(t_rate)} krw</div>", unsafe_allow_html=True)

    t_mode = st.radio("", ["USDT 수량으로 입력", "KRW 금액으로 입력"],
                      horizontal=True, label_visibility="collapsed", key="t_mode")
    if t_mode == "USDT 수량으로 입력":
        tu_amt_raw = extract_int(st.text_input("수량 (USDT)", key="t_u"))
        tu_amt = tu_amt_raw if (t_rate > 0 and tu_amt_raw > 0) else 0
        total_t_krw = tu_amt * t_rate
    else:
        tu_krw_raw = extract_int(st.text_input("금액 (KRW)", key="t_k"))
        if t_rate > 0 and tu_krw_raw > 0:
            total_t_krw = tu_krw_raw
            tu_amt = math.ceil(tu_krw_raw / t_rate)
        else:
            tu_amt = 0; total_t_krw = 0

    if tu_amt > 0 and t_rate > 0:
        my_w = st.session_state.db.get('my_wallet', '')
        t_msg = (f"mid : {selected_m}\ntop-up usdt : {fmt(int(tu_amt))} usdt\n"
                 f"exchange to krw : {fmt(int(total_t_krw))} krw\n1usdt = {fmt(t_rate)} krw\n\n"
                 f"{my_w}\n\nPlease check the invoice and transfer the USDT to the address provided.\n\n"
                 f"⚠️ Please confirm whether you would like to proceed with the top-up.\n"
                 f"Kindly reply as soon as possible.\n"
                 f"Note: The exchange rate provided is based on the current market rate\n"
                 f"and may change if there is a delay in your response.")
        editable_box(t_msg, "green", "res_06_req")

        # 잔액 보고
        bal_after = extract_int(st.text_input("탑업 후 잔액 입력 (KRW)", key="t_bal",
                                               placeholder="탑업 완료 후 머천트 잔액을 입력하세요"))
        if bal_after > 0:
            bal_msg = (f"Top-up completed\n\n"
                       f"Balance update\n\n"
                       f"- {selected_m} : {fmt(bal_after)} krw")
            editable_box(bal_msg, "blue", "res_06_bal")

# ══════════════════════════════════════════════════════════
# 관리자 페이지
# ══════════════════════════════════════════════════════════
elif st.session_state.page == 'admin':
    st.markdown('<div class="main-title">머천트 및 지갑 관리</div>', unsafe_allow_html=True)

    st.markdown("""
    <div style="margin-top:8px;margin-bottom:14px;display:flex;align-items:center;gap:12px;
        padding:10px 16px;background:rgba(255,255,255,0.02);border-radius:8px;
        border-left:3px solid #4a90d9;position:relative;overflow:hidden;">
        <span style="font-family:'Space Mono',monospace;font-size:0.75em;font-weight:700;color:#4a90d9;opacity:0.7;">MY</span>
        <span style="font-family:'Noto Sans KR',sans-serif;font-size:0.92em;font-weight:700;color:#d0dff0;text-transform:uppercase;">내 지갑 주소</span>
    </div>""", unsafe_allow_html=True)
    my_w = st.text_input("내 USDT 지갑 주소", value=st.session_state.db.get('my_wallet', ''), label_visibility="collapsed")
    if st.button("저장", use_container_width=True):
        st.session_state.db['my_wallet'] = my_w
        save_data(st.session_state.db); st.toast("지갑 정보가 저장되었습니다. ✅")

    st.markdown("""
    <div style="margin-top:32px;display:flex;align-items:center;gap:12px;
        padding:10px 16px;background:rgba(255,255,255,0.02);border-radius:8px;
        border-left:3px solid #a855f7;position:relative;overflow:hidden;">
        <span style="font-family:'Space Mono',monospace;font-size:0.75em;font-weight:700;color:#a855f7;opacity:0.7;">NEW</span>
        <span style="font-family:'Noto Sans KR',sans-serif;font-size:0.92em;font-weight:700;color:#d0dff0;text-transform:uppercase;">업체 추가</span>
    </div>""", unsafe_allow_html=True)
    n_name   = st.text_input("업체명", key="n_name")
    n_wallet = st.text_input("지갑주소", key="n_wallet")
    n_fee    = st.text_input("마크업 수수료 (%)", value="0.5", key="n_fee")
    n_note   = st.text_input("비고", key="n_note")
    if st.button("등록", use_container_width=True, key="btn_register"):
        if n_name:
            st.session_state.db['merchants'][n_name] = {"wallet": n_wallet, "fee": n_fee, "note": n_note}
            save_data(st.session_state.db)
            st.rerun()

    st.markdown("""
    <div style="margin-top:32px;margin-bottom:14px;display:flex;align-items:center;gap:12px;
        padding:10px 16px;background:rgba(255,255,255,0.02);border-radius:8px;
        border-left:3px solid #2ecc71;position:relative;overflow:hidden;">
        <span style="font-family:'Space Mono',monospace;font-size:0.75em;font-weight:700;color:#2ecc71;opacity:0.7;">LIST</span>
        <span style="font-family:'Noto Sans KR',sans-serif;font-size:0.92em;font-weight:700;color:#d0dff0;text-transform:uppercase;">등록 업체 관리</span>
    </div>""", unsafe_allow_html=True)
    for name in sorted(st.session_state.db['merchants'].keys()):
        with st.expander(f"📦 {name} 관리"):
            info = st.session_state.db['merchants'][name]
            u_w = st.text_input("지갑주소",          value=info['wallet'],        key=f"w_{name}")
            u_f = st.text_input("마크업 수수료 (%)", value=info['fee'],           key=f"f_{name}")
            u_n = st.text_input("비고",             value=info.get('note',''),   key=f"n_{name}")
            col_save, col_del = st.columns([3, 1])
            with col_save:
                if st.button("변경사항 저장", key=f"s_{name}", use_container_width=True):
                    st.session_state.db['merchants'][name] = {"wallet": u_w, "fee": u_f, "note": u_n}
                    save_data(st.session_state.db); st.toast(f"{name} 저장됨 ✅")
            with col_del:
                if st.button("삭제", key=f"d_{name}", use_container_width=True):
                    del st.session_state.db['merchants'][name]
                    save_data(st.session_state.db); st.rerun()

# ══════════════════════════════════════════════════════════
# 수수료 정산 페이지
# ══════════════════════════════════════════════════════════
elif st.session_state.page == 'commission':
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    st.markdown('<div class="main-title">수수료 정산</div>', unsafe_allow_html=True)

    DRAGON_PARTNERS = ['dr188', 'drbetssen', 'drSpinmama', 'Dpinnacle']

    section_header("01", "내역 텍스트 입력", "#4a90d9", "74,144,217")
    raw_commission = st.text_area("📋 장부 텍스트 붙여넣기", height=250, key="comm_raw",
                                   placeholder="날짜\\t금액\\t환율1\\t환율2\\tUSDT\\t손익\\t메모 형식으로 붙여넣으세요")

    col_d1, col_d2 = st.columns(2)
    with col_d1: date_from = st.text_input("시작일 (예: 3/18/2026)", key="comm_from")
    with col_d2: date_to   = st.text_input("종료일 (예: 4/3/2026)",  key="comm_to")

    if raw_commission:
        topup_records   = []
        settle_records  = []

        for line in raw_commission.strip().split('\n'):
            cols = line.split('\t')
            if len(cols) < 7: continue
            date = cols[0].strip()
            memo = cols[6].strip()

            krw_col  = cols[1].strip().replace(',', '')
            krw_amt  = int(float(krw_col)) if krw_col not in ('무','') else 0

            rate_col = cols[3].strip().replace(',', '')
            rate_col2= cols[2].strip().replace(',', '')
            rate     = float(rate_col) if rate_col not in ('무','') else (float(rate_col2) if rate_col2 not in ('무','') else 0)

            usdt_col = cols[4].strip().replace(',', '')
            usdt_qty = int(float(usdt_col)) if usdt_col not in ('무','') else 0

            for p in DRAGON_PARTNERS:
                if p.lower() in memo.lower():
                    if '탑업' in memo:
                        topup_records.append({'date': date, 'partner': p, 'usdt': usdt_qty, 'krw': krw_amt, 'rate': rate})
                    elif '업체정산' in memo and '에이전트' not in memo and '게이트' not in memo:
                        settle_records.append({'date': date, 'partner': p, 'usdt': usdt_qty, 'krw': krw_amt, 'rate': rate})
                    break

        # ── 결과 표시 ──
        import pandas as pd
        from collections import defaultdict

        section_header("02", "수수료 내역", "#4a90d9", "74,144,217")

        topup_by  = defaultdict(list)
        settle_by = defaultdict(list)
        for r in topup_records:  topup_by[r['partner']].append(r)
        for r in settle_records: settle_by[r['partner']].append(r)

        any_data = False
        for p in DRAGON_PARTNERS:
            t_recs = topup_by.get(p, [])
            s_recs = settle_by.get(p, [])
            if not t_recs and not s_recs: continue
            any_data = True

            st.markdown(f"### ▶ {p}")

            # 탑업
            if t_recs:
                total_usdt = sum(r['usdt'] for r in t_recs)
                total_krw_t = sum(r['krw'] for r in t_recs)
                fee_usdt = round(total_usdt * 0.005, 2)
                df_t = pd.DataFrame([{'날짜': r['date'], 'USDT': f"{r['usdt']:,}", '환율': r['rate'], 'KRW': f"{r['krw']:,}"} for r in t_recs])
                df_t.loc['합계'] = ['합계', f"{total_usdt:,}", '', f"{total_krw_t:,}"]
                st.caption("📤 탑업 내역")
                st.dataframe(df_t, use_container_width=True, hide_index=True)
                fee_krw_t = round(total_krw_t * 0.005)
                st.success(f"수수료 (0.5%) : **{fee_usdt:,.2f} usdt / {fee_krw_t:,} krw**")

            # 정산
            if s_recs:
                total_krw_s = sum(r['krw'] for r in s_recs)
                fee_krw = round(total_krw_s * 0.005)
                df_s = pd.DataFrame([{'날짜': r['date'], '환율': r['rate'], 'KRW': f"{r['krw']:,}"} for r in s_recs])
                df_s.loc['합계'] = ['합계', '', f"{total_krw_s:,}"]
                st.caption("💱 정산 내역")
                st.dataframe(df_s, use_container_width=True, hide_index=True)
                st.success(f"정산 수수료 (0.5%) : **{fee_krw:,} krw**")

            st.divider()

        if not any_data:
            st.info("해당 기간 드래곤 파트너 내역 없음")

        # ── 엑셀 생성 ──
        section_header("04", "엑셀 다운로드", "#a855f7", "168,85,247")

        def make_excel(topup_records, settle_records):
            wb = Workbook()

            # 스타일
            hdr_font    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            hdr_fill    = PatternFill("solid", start_color="1E3A5F")
            sub_fill    = PatternFill("solid", start_color="2E4A7F")
            total_fill  = PatternFill("solid", start_color="0F2040")
            total_font  = Font(bold=True, color="38BDF8", name="Arial", size=10)
            center      = Alignment(horizontal="center", vertical="center")
            right       = Alignment(horizontal="right",  vertical="center")
            thin        = Side(style="thin", color="334155")
            border      = Border(left=thin, right=thin, top=thin, bottom=thin)
            num_fmt     = '#,##0'

            from collections import defaultdict

            # ── 탑업 시트 ──
            ws1 = wb.active
            ws1.title = "탑업 내역"
            ws1.append(["Date", "Partner", "USDT amount", "Rate", "KRW amount"])
            for cell in ws1[1]:
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = center; cell.border = border

            topup_by = defaultdict(list)
            for r in topup_records: topup_by[r['partner']].append(r)

            data_rows = []
            for p in DRAGON_PARTNERS:
                for r in topup_by.get(p, []):
                    ws1.append([r['date'], r['partner'], r['usdt'], r['rate'], r['krw']])
                    data_rows.append(ws1.max_row)

            # 파트너별 소계
            for p in DRAGON_PARTNERS:
                recs = topup_by.get(p, [])
                if not recs: continue
                total_u = sum(r['usdt'] for r in recs)
                total_k = sum(r['krw']  for r in recs)
                fee_u   = round(total_u * 0.005, 2)
                ws1.append([f"{p} 소계", "", total_u, "", total_k])
                row = ws1.max_row
                for cell in ws1[row]:
                    cell.font = total_font; cell.fill = total_fill; cell.border = border
                ws1.append([f"수수료 0.5%", "", fee_u, "", ""])
                row2 = ws1.max_row
                for cell in ws1[row2]:
                    cell.font = Font(bold=True, color="2ECC71", name="Arial", size=10)
                    cell.fill = PatternFill("solid", start_color="0A2016")
                    cell.border = border

            # 숫자 포맷
            for row in ws1.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    if isinstance(cell.value, (int, float)) and cell.column in (3, 5):
                        cell.number_format = num_fmt

            ws1.column_dimensions['A'].width = 14
            ws1.column_dimensions['B'].width = 14
            ws1.column_dimensions['C'].width = 14
            ws1.column_dimensions['D'].width = 10
            ws1.column_dimensions['E'].width = 16

            # ── 정산 시트 ──
            ws2 = wb.create_sheet("업체정산 내역")
            ws2.append(["Date", "Partner", "Rate", "KRW amount"])
            for cell in ws2[1]:
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = center; cell.border = border

            settle_by = defaultdict(list)
            for r in settle_records: settle_by[r['partner']].append(r)

            for p in DRAGON_PARTNERS:
                for r in settle_by.get(p, []):
                    ws2.append([r['date'], r['partner'], r['rate'], r['krw']])

            for p in DRAGON_PARTNERS:
                recs = settle_by.get(p, [])
                if not recs: continue
                total_k = sum(r['krw'] for r in recs)
                fee_k   = round(total_k * 0.005)
                ws2.append([f"{p} 소계", "", "", total_k])
                row = ws2.max_row
                for cell in ws2[row]:
                    cell.font = total_font; cell.fill = total_fill; cell.border = border
                ws2.append([f"수수료 0.5%", "", "", fee_k])
                row2 = ws2.max_row
                for cell in ws2[row2]:
                    cell.font = Font(bold=True, color="F39C12", name="Arial", size=10)
                    cell.fill = PatternFill("solid", start_color="241800")
                    cell.border = border

            for row in ws2.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    if isinstance(cell.value, (int, float)) and cell.column == 4:
                        cell.number_format = num_fmt

            ws2.column_dimensions['A'].width = 14
            ws2.column_dimensions['B'].width = 14
            ws2.column_dimensions['C'].width = 10
            ws2.column_dimensions['D'].width = 16

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.getvalue()

        excel_data = make_excel(topup_records, settle_records)
        period = f"{date_from}~{date_to}".replace('/', '') if date_from and date_to else "수수료정산"
        st.download_button(
            label="📥 엑셀 다운로드",
            data=excel_data,
            file_name=f"드래곤수수료_{period}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# ══════════════════════════════════════════════════════════
# 에이전트 정산 페이지
# ══════════════════════════════════════════════════════════
elif st.session_state.page == 'agent':
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    st.markdown('<div class="main-title">에이전트 정산</div>', unsafe_allow_html=True)

    # 에이전트 설정
    AGENTS = {
        'Dean': {
            'merchants': {
                'dr188':    {'name': 'dr188',     'dep_rate': 0.001, 'wds_rate': 0.001, 'gate_dep': 0.012, 'gate_wds': 0.007},
                'drbetssen':{'name': 'drBetssen', 'dep_rate': 0.001, 'wds_rate': 0.001, 'gate_dep': 0.014, 'gate_wds': 0.006},
            }
        },
        'Tofi': {
            'merchants': {
                'spfxm':    {'name': 'spfxm', 'dep_rate': 0.002, 'wds_rate': 0.001, 'gate_dep': 0, 'gate_wds': 0},
            }
        },
        'Michell': {
            'merchants': {
                'v99_BT':   {'name': 'v99_BT', 'dep_rate': 0.001, 'wds_rate': 0.001, 'gate_dep': 0, 'gate_wds': 0},
            }
        },
    }

    def parse_mbd_summary(text):
        """Merchant By Date Summary 줄에서 입금/출금 합계 추출"""
        summary_m = re.search(r'Summary([\d,]+)', text)
        if not summary_m: return 0, 0
        nums_raw = re.findall(r'[\d,]+', text[text.find('Summary'):])
        nums = [int(n.replace(',','')) for n in nums_raw if n.replace(',','').isdigit()]
        if len(nums) >= 2:
            return nums[0], nums[2] if len(nums) > 2 else 0
        return 0, 0

    def make_agent_excel(agent_name, results, period):
        wb = Workbook()
        # 스타일
        title_font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
        title_fill = PatternFill("solid", start_color="4A7AB5")
        hdr1_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        hdr1_fill  = PatternFill("solid", start_color="6A9FD4")
        hdr2_font  = Font(bold=True, color="2C3E50", name="Arial", size=10)
        hdr2_fill  = PatternFill("solid", start_color="BDD7EE")
        data_font  = Font(name="Arial", size=10, color="2C3E50")
        data_fill  = PatternFill("solid", start_color="FFFFFF")
        total_font = Font(bold=True, color="1A4A7A", name="Arial", size=10)
        total_fill = PatternFill("solid", start_color="DDEEFF")
        fee_font   = Font(bold=True, color="1A6B1A", name="Arial", size=10)
        fee_fill   = PatternFill("solid", start_color="C6EFCE")
        thin       = Side(style="thin", color="9DC3E6")
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)
        center     = Alignment(horizontal="center", vertical="center")
        right      = Alignment(horizontal="right", vertical="center")
        num_fmt    = '#,##0'
        pct_fmt    = '0.00%'

        ws = wb.active
        ws.title = "Agent Settlement"

        # 행1: 타이틀 (에이전트명 + 기간)
        ws.merge_cells('A1:G1')
        ws['A1'] = f"{agent_name} Agent Settlement ({period})"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = center

        # 행2: 대헤더
        ws.append(["", "Deposits", "", "", "Withdrawals", "", "Total"])
        for cell in ws[2]:
            cell.font = hdr1_font; cell.fill = hdr1_fill
            cell.alignment = center; cell.border = border
        ws.merge_cells('B2:D2')
        ws.merge_cells('E2:F2')

        # 행3: 소헤더
        ws.append(["Merchant", "Rate", "KRW", "Amount", "Rate", "KRW", "Amount"])
        for cell in ws[3]:
            cell.font = hdr2_font; cell.fill = hdr2_fill
            cell.alignment = center; cell.border = border

        total_dep_fee = 0
        total_wds_fee = 0

        gate_fill  = PatternFill("solid", start_color="FFF2CC")
        gate_font  = Font(bold=True, color="7B5C00", name="Arial", size=10)
        total_gate = 0

        # ── 시트1: 에이전트 정산 ──
        for mid, res in results.items():
            cfg = AGENTS[agent_name]['merchants'][mid]
            dep_fee      = res['dep_fee']
            wds_fee      = res['wds_fee']
            gate_dep_fee = round(res['deposits']    * cfg.get('gate_dep', 0))
            gate_wds_fee = round(res['withdrawals'] * cfg.get('gate_wds', 0))
            total_dep_fee += dep_fee
            total_wds_fee += wds_fee
            total_gate    += gate_dep_fee + gate_wds_fee

            ws.append([cfg['name'], cfg['dep_rate'], res['deposits'], dep_fee,
                       cfg['wds_rate'], res['withdrawals'], wds_fee])
            r = ws.max_row
            for col in range(1, 8):
                cell = ws.cell(r, col)
                cell.font = data_font; cell.fill = data_fill
                cell.border = border; cell.alignment = center
                if col in (2, 5): cell.number_format = pct_fmt
                if col in (3, 4, 6, 7): cell.number_format = num_fmt

        grand_total = total_dep_fee + total_wds_fee
        ws.append(["Subtotal", "", "", total_dep_fee, "", "", total_wds_fee])
        r = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(r, col)
            cell.font = total_font; cell.fill = total_fill
            cell.border = border; cell.alignment = center
            if col in (4, 7): cell.number_format = num_fmt

        ws.append(["Total Fee", "", "", "", "", "", grand_total])
        r = ws.max_row
        ws.merge_cells(f'A{r}:F{r}')
        for col in range(1, 8):
            cell = ws.cell(r, col)
            cell.font = fee_font; cell.fill = fee_fill
            cell.border = border; cell.alignment = center
            if col == 7: cell.number_format = num_fmt

        # ── 시트2: 게이트 정산 ──
        if total_gate > 0:
            ws2 = wb.create_sheet("Gate Settlement")
            ws2.merge_cells('A1:G1')
            ws2['A1'] = f"dragong Gate Settlement ({period})"
            ws2['A1'].font = title_font
            ws2['A1'].fill = title_fill
            ws2['A1'].alignment = center

            ws2.append(["", "Deposits", "", "", "Withdrawals", "", "Total"])
            for cell in ws2[2]:
                cell.font = hdr1_font; cell.fill = hdr1_fill
                cell.alignment = center; cell.border = border
            ws2.merge_cells('B2:D2')
            ws2.merge_cells('E2:F2')

            ws2.append(["Merchant", "Rate", "KRW", "Amount", "Rate", "KRW", "Amount"])
            for cell in ws2[3]:
                cell.font = hdr2_font; cell.fill = hdr2_fill
                cell.alignment = center; cell.border = border

            total_gate_dep = 0
            total_gate_wds = 0
            for mid, res in results.items():
                cfg = AGENTS[agent_name]['merchants'][mid]
                if cfg.get('gate_dep', 0) == 0 and cfg.get('gate_wds', 0) == 0:
                    continue
                gate_dep_fee = round(res['deposits']    * cfg.get('gate_dep', 0))
                gate_wds_fee = round(res['withdrawals'] * cfg.get('gate_wds', 0))
                total_gate_dep += gate_dep_fee
                total_gate_wds += gate_wds_fee

                ws2.append([cfg['name'], cfg.get('gate_dep', 0), res['deposits'], gate_dep_fee,
                            cfg.get('gate_wds', 0), res['withdrawals'], gate_wds_fee])
                r2 = ws2.max_row
                for col in range(1, 8):
                    cell = ws2.cell(r2, col)
                    cell.font = data_font; cell.fill = data_fill
                    cell.border = border; cell.alignment = center
                    if col in (2, 5): cell.number_format = pct_fmt
                    if col in (3, 4, 6, 7): cell.number_format = num_fmt

            gate_grand = total_gate_dep + total_gate_wds
            ws2.append(["Subtotal", "", "", total_gate_dep, "", "", total_gate_wds])
            r2 = ws2.max_row
            for col in range(1, 8):
                cell = ws2.cell(r2, col)
                cell.font = total_font; cell.fill = total_fill
                cell.border = border; cell.alignment = center
                if col in (4, 7): cell.number_format = num_fmt

            ws2.append(["Total Gate Fee", "", "", "", "", "", gate_grand])
            r2 = ws2.max_row
            ws2.merge_cells(f'A{r2}:F{r2}')
            for col in range(1, 8):
                cell = ws2.cell(r2, col)
                cell.font = fee_font; cell.fill = fee_fill
                cell.border = border; cell.alignment = center
                if col == 7: cell.number_format = num_fmt

            for col, w in zip('ABCDEFG', [14, 8, 16, 12, 8, 16, 12]):
                ws2.column_dimensions[col].width = w
            ws2.row_dimensions[1].height = 22
            ws2.row_dimensions[2].height = 18

        # 컬럼 너비
        for col, w in zip('ABCDEFG', [14, 8, 16, 12, 8, 16, 12]):
            ws.column_dimensions[col].width = w
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    # ── UI ──
    selected_agent = st.radio("", list(AGENTS.keys()),
                               horizontal=True, label_visibility="collapsed", key="agent_sel")

    st.divider()
    section_header("01", f"{selected_agent} 머천트 입력", "#4a90d9", "74,144,217")

    col_d1, col_d2 = st.columns(2)
    with col_d1:
        agent_date_from_d = st.date_input("시작일", value=None, key="agent_from_d", format="YYYY-MM-DD")
        agent_date_from = agent_date_from_d.strftime("%Y-%m-%d") if agent_date_from_d else ""
    with col_d2:
        agent_date_to_d = st.date_input("종료일", value=None, key="agent_to_d", format="YYYY-MM-DD")
        agent_date_to = agent_date_to_d.strftime("%Y-%m-%d") if agent_date_to_d else ""
    if selected_agent != 'Michell' and agent_date_from and agent_date_to:
        st.caption("⚠️ 어드민에서 해당 기간 설정 후 긁어온 데이터를 붙여넣으세요. 날짜는 엑셀 타이틀 표시용입니다.")

    merchant_inputs = {}
    for mid, cfg in AGENTS[selected_agent]['merchants'].items():
        st.text(f"📋 {cfg['name']} ({mid}) 내역")
        if selected_agent == 'Michell':
            merchant_inputs[mid] = st.text_area(
                f"{cfg['name']} 붙여넣기",
                height=120, key=f"agent_{mid}",
                placeholder="ADA 어드민 일일통계 페이지 전체를 붙여넣으세요\n(날짜 필터는 아래 시작일/종료일로 적용됩니다)",
                label_visibility="collapsed"
            )
        else:
            merchant_inputs[mid] = st.text_area(
                f"{cfg['name']} 붙여넣기",
                height=120, key=f"agent_{mid}",
                placeholder="Merchant By Date Statistics 페이지 전체를 붙여넣으세요",
                label_visibility="collapsed"
            )

    if st.button("📊 정산 계산", use_container_width=True, key="agent_calc"):
        st.session_state['agent_results'] = {}
        st.session_state['agent_name'] = selected_agent
        for mid, text in merchant_inputs.items():
            if not text.strip(): continue
            cfg = AGENTS[selected_agent]['merchants'][mid]

            if selected_agent == 'Michell':
                # ADA 어드민 일일통계 파싱
                # 구조: 날짜 한줄 → 입금 → 입금수수료 → 출금 → 출금수수료 → 게이트 → 에이전트 → 본사순이익
                deposits = 0
                withdrawals = 0
                from datetime import datetime as dt
                # 붙여넣기 시 날짜가 헤더에 붙어오는 경우 분리
                text_clean = re.sub(r'(\d{4}-\d{2}-\d{2})', r'\n\1', text)
                lines = [l.strip() for l in text_clean.strip().split('\n') if l.strip()]
                i = 0
                while i < len(lines):
                    date_m = re.match(r'(\d{4}-\d{2}-\d{2})', lines[i])
                    if date_m:
                        row_date = date_m.group(1)
                        # 날짜 필터
                        include = True
                        if agent_date_from:
                            try:
                                if dt.strptime(row_date, '%Y-%m-%d') < dt.strptime(agent_date_from, '%Y-%m-%d'):
                                    include = False
                            except: pass
                        if agent_date_to:
                            try:
                                if dt.strptime(row_date, '%Y-%m-%d') > dt.strptime(agent_date_to, '%Y-%m-%d'):
                                    include = False
                            except: pass
                        # 날짜 다음 7개 줄에서 숫자 추출
                        nums = []
                        j = i + 1
                        while j < len(lines) and len(nums) < 7:
                            clean = lines[j].replace(',', '').replace('₩', '').strip()
                            if re.match(r'^\d+$', clean):
                                nums.append(int(clean))
                            elif re.match(r'\d{4}-\d{2}-\d{2}', lines[j]):
                                break
                            j += 1
                        if include and len(nums) >= 4:
                            deposits    += nums[0]  # 입금
                            withdrawals += nums[2]  # 출금
                        i = j
                    else:
                        i += 1
                dep_fee = round(deposits    * cfg['dep_rate'])
                wds_fee = round(withdrawals * cfg['wds_rate'])
            else:
                # TL MBD 파싱 (Summary 줄에서 합계 추출)
                summary_idx = text.find('Summary')
                if summary_idx == -1: continue
                summary_part = text[summary_idx:]
                nums = [int(n.replace(',','')) for n in re.findall(r'[\d,]+', summary_part) if n.replace(',','').isdigit()]
                if len(nums) < 3: continue
                deposits    = nums[0]
                withdrawals = nums[2]
                dep_fee     = round(deposits    * cfg['dep_rate'])
                wds_fee     = round(withdrawals * cfg['wds_rate'])

            st.session_state['agent_results'][mid] = {
                'deposits': deposits, 'withdrawals': withdrawals,
                'dep_fee': dep_fee,   'wds_fee': wds_fee
            }

    if st.session_state.get('agent_results'):
        results = st.session_state['agent_results']
        section_header("02", "정산 결과", "#2ecc71", "46,204,113")

        total_dep_fee = 0
        total_wds_fee = 0

        total_gate_fee = 0
        for mid, res in results.items():
            cfg = AGENTS[selected_agent]['merchants'][mid]
            gate_dep_fee = round(res['deposits']    * cfg.get('gate_dep', 0))
            gate_wds_fee = round(res['withdrawals'] * cfg.get('gate_wds', 0))
            gate_total   = gate_dep_fee + gate_wds_fee
            total_gate_fee += gate_total

            st.markdown(f"**▶ {cfg['name']} ({mid})**")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("입금", f"{res['deposits']:,} krw")
                st.metric(f"입금 수수료 ({cfg['dep_rate']*100}%)", f"{res['dep_fee']:,} krw")
            with col2:
                st.metric("출금", f"{res['withdrawals']:,} krw")
                st.metric(f"출금 수수료 ({cfg['wds_rate']*100}%)", f"{res['wds_fee']:,} krw")
            with col3:
                st.metric("에이전트 합계", f"{res['dep_fee'] + res['wds_fee']:,} krw")
            if gate_total > 0:
                gc1, gc2, gc3 = st.columns(3)
                with gc1: st.metric(f"게이트 입금 ({cfg['gate_dep']*100:.1f}%)", f"{gate_dep_fee:,} krw")
                with gc2: st.metric(f"게이트 출금 ({cfg['gate_wds']*100:.1f}%)", f"{gate_wds_fee:,} krw")
                with gc3: st.metric("게이트 합계", f"{gate_total:,} krw")
            total_dep_fee += res['dep_fee']
            total_wds_fee += res['wds_fee']
            st.divider()

        grand_total = total_dep_fee + total_wds_fee
        st.success(f"**{selected_agent} 에이전트 수수료 : {grand_total:,} krw**")
        if total_gate_fee > 0:
            st.info(f"**{selected_agent} 게이트 수수료 : {total_gate_fee:,} krw**")

        # 엑셀 다운로드
        section_header("03", "엑셀 다운로드", "#a855f7", "168,85,247")
        period_file = f"{agent_date_from}~{agent_date_to}".replace('/','') if agent_date_from and agent_date_to else "에이전트정산"
        period_title = f"{agent_date_from}~{agent_date_to}" if agent_date_from and agent_date_to else ""
        excel_data = make_agent_excel(selected_agent, results, period_title)
        st.download_button(
            label="📥 엑셀 다운로드",
            data=excel_data,
            file_name=f"{selected_agent}_에이전트정산_{period_file}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )